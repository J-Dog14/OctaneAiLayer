"""
Parse the 8ctane "Lift Template progression.xlsx" into a structured JSON
catalog at references/template_catalog.json.

The spreadsheet has two sheets:

  Template Progression: ~10k rows of template blocks. Each block is led by a
    "header" row whose column G contains a template ID like '4222 - 11' or
    'Baseball Club Total Body - 1 - 1', and whose columns J/M/P may hold
    higher-progression-level header IDs (L2/L3/L4). The exercise rows that
    follow (until the next header or two blank rows) contain one exercise per
    row with sets×reps in columns G/H (L1), J/K (L2), M/N (L3), P/Q (L4).

  Index: 1009 rows mapping template_id → description / "Made For" athlete /
    notes. This is the gold link between a template and historical use.

Run:
    python -m scripts.parse_templates path/to/Lift\ Template\ progression.xlsx

Output: ../references/template_catalog.json (next to this script's parent).
"""
from __future__ import annotations

import json
import re
import sys
from collections import defaultdict
from pathlib import Path

import openpyxl

# ID like "4222 - 11" or "4222-11" or "4222 - 11 - Description"
_ID_RX = re.compile(r"^\s*(?P<num>\d{4,7})\s*[\-–]\s*(?P<meso>\d{1,3})\b")

# Named header pattern like "Baseball Club Total Body - 1 - 1"
_NAMED_RX = re.compile(
    r"^(?P<desc>[A-Za-z][^-]*?)\s*[\-–]\s*(?P<meso>\d{1,2})\s*[\-–]\s*\d{1,2}\s*$"
)

# Progression-level column positions (0-indexed in row tuple).
# Each level uses two columns: (exercise_name, sets_x_reps).
_LEVELS = [
    ("L1", 6,  7),   # G, H
    ("L2", 9,  10),  # J, K
    ("L3", 12, 13),  # M, N
    ("L4", 15, 16),  # P, Q
]


def _is_header_row(g: str) -> bool:
    if not g:
        return False
    s = str(g).strip()
    return bool(_ID_RX.match(s) or _NAMED_RX.match(s))


def _parse_header_cell(s: str) -> dict | None:
    """Pull out template_id, mesocycle, and optional description from a header cell."""
    if not s:
        return None
    s = str(s).strip()
    m = _ID_RX.match(s)
    if m:
        rest = s[m.end():].lstrip(" -–")
        return {
            "raw": s,
            "id_numeric": m.group("num"),
            "mesocycle_id": m.group("meso"),
            "description": rest or None,
        }
    m = _NAMED_RX.match(s)
    if m:
        return {
            "raw": s,
            "id_numeric": None,
            "mesocycle_id": m.group("meso"),
            "description": m.group("desc").strip(),
        }
    return None


def _decode_id(numeric_id: str | None) -> dict:
    """Decode the 4-7 digit ID into its semantic components."""
    if not numeric_id or not numeric_id.isdigit():
        return {}
    digits = list(numeric_id)
    LEVEL = {"1": "Before Baseball Club (Beginner)", "2": "Baseball Club (Advanced)",
             "3": "High School (Beginner)", "4": "High School", "5": "College",
             "6": "Pro", "7": "Softball", "8": "Misc"}
    FOCUS = {"1": "Strength", "2": "Power", "3": "Speed", "4": "In-Season",
             "5": "Hypertrophy"}
    MOVEMENT = {"1": "Legs", "2": "Upper", "3": "Total Body", "4": "Sprint",
                "5": "Jump", "6": "Push", "7": "Pull", "8": "Core", "9": "Recovery"}
    out = {}
    if len(digits) >= 1: out["level"] = LEVEL.get(digits[0], "?")
    if len(digits) >= 2: out["focus"] = FOCUS.get(digits[1], "?")
    if len(digits) >= 3: out["sprint_days_per_week"] = int(digits[2]) if digits[2].isdigit() else None
    if len(digits) >= 4: out["identifier"] = digits[3]
    if len(digits) >= 5: out["mesocycle_number"] = digits[4]
    if len(digits) >= 6: out["movement_bucket"] = MOVEMENT.get(digits[5], "?")
    if len(digits) >= 7: out["lift_position_in_day"] = digits[6]
    return out


def _parse_template_block(rows: list[tuple]) -> dict:
    """Given the header row + exercise rows (until the next blank/header),
    extract the 4 levels of progressions."""
    header = rows[0]
    exercise_rows = rows[1:]

    block = {"levels": {}}
    # Header parse for each level slot
    for tag, name_col, _ in _LEVELS:
        h = _parse_header_cell(header[name_col] if name_col < len(header) else None)
        if h:
            block["levels"][tag] = {
                "header": h,
                "decoded_id": _decode_id(h.get("id_numeric")),
                "exercises": [],
            }

    for row in exercise_rows:
        for tag, name_col, sxr_col in _LEVELS:
            if tag not in block["levels"]:
                continue
            name = row[name_col] if name_col < len(row) else None
            sxr  = row[sxr_col]  if sxr_col  < len(row) else None
            if name is None or str(name).strip() == "" or str(name).strip() == "0":
                continue
            block["levels"][tag]["exercises"].append({
                "exercise": str(name).strip(),
                "sets_x_reps": (str(sxr).strip() if sxr is not None else None),
            })
    return block


def parse_workbook(path: Path) -> dict:
    wb = openpyxl.load_workbook(path, data_only=True)

    # 1. Index sheet → template_id → metadata
    index_ws = wb["Index"]
    index: dict[str, dict] = {}
    for row in index_ws.iter_rows(values_only=True, min_row=2):
        if not row or row[0] is None:
            continue
        tid_raw = row[0]
        tid = str(tid_raw).replace(".0", "").strip()
        desc = (row[1] if len(row) > 1 else None) or None
        made_for = (row[2] if len(row) > 2 else None) or None
        notes = (row[3] if len(row) > 3 else None) or None
        index.setdefault(tid, {
            "template_id_numeric": tid,
            "description": desc and str(desc).strip(),
            "made_for_athletes": [],
            "notes": notes and str(notes).strip(),
            "decoded": _decode_id(tid),
        })
        if made_for:
            index[tid]["made_for_athletes"].append(str(made_for).strip())

    # 2. Template Progression sheet — collect blocks
    ws = wb["Template Progression"]
    rows = [row for row in ws.iter_rows(values_only=True)]
    # ensure each row is at least 17 cols
    rows = [(r + (None,) * 17)[:32] for r in rows]

    blocks: list[dict] = []
    i = 0
    while i < len(rows):
        row = rows[i]
        g = row[6]
        if g is not None and _is_header_row(g):
            # Collect rows until next header or 2 consecutive empty (G & J & M & P all empty)
            block_rows = [row]
            j = i + 1
            consecutive_empty = 0
            while j < len(rows):
                nxt = rows[j]
                if _is_header_row(nxt[6]):
                    break
                # Empty check
                key_cells = [nxt[c] for c in (6, 9, 12, 15)]
                if all(c is None or str(c).strip() in ("", "0") for c in key_cells):
                    consecutive_empty += 1
                    if consecutive_empty >= 2:
                        break
                else:
                    consecutive_empty = 0
                block_rows.append(nxt)
                j += 1
            blk = _parse_template_block(block_rows)
            blk["source_row_start"] = i + 1
            blocks.append(blk)
            i = j
        else:
            i += 1

    # 3. Attach index metadata to each block's L1
    enriched = []
    for blk in blocks:
        # Use L1 numeric id as the canonical lookup
        l1 = blk["levels"].get("L1") or {}
        h = l1.get("header") or {}
        tid = h.get("id_numeric")
        if tid and tid in index:
            blk["index_metadata"] = index[tid]
        enriched.append(blk)

    # 4. Aggregate templates by id_numeric (multiple blocks may share an ID prefix
    #    if they're different days/positions like 4111-11, 4111-12, 4111-13...)
    templates_by_id = defaultdict(list)
    for blk in enriched:
        tid = ((blk["levels"].get("L1") or {}).get("header") or {}).get("id_numeric")
        if tid:
            templates_by_id[tid].append(blk)

    return {
        "n_blocks_parsed": len(enriched),
        "n_unique_template_ids": len(templates_by_id),
        "index": index,
        "blocks": enriched,
        "templates_by_id": templates_by_id,
    }


def main() -> int:
    if len(sys.argv) < 2:
        print("Usage: python -m scripts.parse_templates <path/to/spreadsheet.xlsx>")
        return 2
    src = Path(sys.argv[1])
    if not src.exists():
        print(f"File not found: {src}")
        return 2
    out_dir = Path(__file__).resolve().parent.parent / "references"
    out_dir.mkdir(parents=True, exist_ok=True)
    out = out_dir / "template_catalog.json"

    catalog = parse_workbook(src)
    out.write_text(json.dumps(catalog, indent=2, default=str), encoding="utf-8")
    print(f"Wrote {out}")
    print(f"  blocks parsed: {catalog['n_blocks_parsed']}")
    print(f"  unique template IDs: {catalog['n_unique_template_ids']}")
    print(f"  athletes assigned in index: "
          f"{sum(len(v['made_for_athletes']) for v in catalog['index'].values())}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
